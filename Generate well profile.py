"""
generate_well_profile.py
Generates a ready-to-import Excel file for WellLogAnalyzer.

Sheet layout matches ImportFromExcelButton.parseExcel() exactly:
  - WellInfo   : row 2 = WellName, TotalDepth, CasingOD, CasingID
  - DrillString: (not parsed by current importer — filled for reference only)
  - FluidProps : row 2 = MudWeight, FlowRate, PV, YP
  - Formations : rows 2+ = ZoneName, TopDepth, BottomDepth, PP, FG, Lithology
  - Survey     : rows 2+ = MD, Inclination, Azimuth

Run:
    pip install openpyxl
    python generate_well_profile.py
Output: Well_07A_Profile.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ── Colour palette (matches the app's dark theme accents) ────────────────────
NAVY      = "0D1B2A"
SLATE     = "1A2535"
CARD      = "212E42"
AMBER     = "F4A917"
TEAL      = "2EC4B6"
WHITE     = "FFFFFF"
GREY_TEXT = "94A3B8"
CORAL     = "E63946"

def header_fill():
    return PatternFill("solid", fgColor=CARD)

def amber_fill():
    return PatternFill("solid", fgColor=AMBER)

def row_fill(even: bool):
    return PatternFill("solid", fgColor=SLATE if even else NAVY)

def header_font():
    return Font(bold=True, color=AMBER, name="Calibri", size=10)

def value_font(color=WHITE):
    return Font(color=color, name="Calibri", size=10)

def thin_border():
    s = Side(style="thin", color="2E3E58")
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center")

def write_header_row(ws, headers: list, row: int = 1):
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=title)
        cell.fill    = header_fill()
        cell.font    = header_font()
        cell.border  = thin_border()
        cell.alignment = center()
        ws.column_dimensions[get_column_letter(col)].width = max(18, len(title) + 4)

def write_data_row(ws, values: list, row: int, even: bool = True):
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill      = row_fill(even)
        cell.font      = value_font()
        cell.border    = thin_border()
        cell.alignment = center()

def style_sheet(ws):
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = AMBER

# ════════════════════════════════════════════════════════════════════════════
# DATA — edit these values to match the well you want to simulate
# ════════════════════════════════════════════════════════════════════════════

WELL_INFO = {
    "WellName"   : "Well-07A",
    "TotalDepth" : 9500.0,      # ft
    "CasingOD"   : 9.625,       # in  (9-5/8" surface casing)
    "CasingID"   : 8.835,       # in
}

# Drill string geometry (informational — not parsed by the current importer)
DRILL_STRING = {
    "DrillPipeOD"      : 5.0,    # in
    "DrillPipeID"      : 4.276,  # in
    "DrillCollarOD"    : 6.5,    # in
    "DrillCollarLength": 600.0,  # ft
}

# Individual drill string sections (also informational)
DRILL_SECTIONS = [
    # Name,            OD(in), ID(in), Length(ft), Weight(lb/ft)
    ("Drill Pipe",     5.0,    4.276,  8300.0,     19.5),
    ("Drill Collar",   6.5,    2.813,   600.0,     83.0),
    ("HWDP",           5.0,    3.0,     600.0,     50.0),
]

FLUID_PROPS = {
    "MudWeight" : 10.5,   # ppg
    "FlowRate"  : 400.0,  # gpm
    "PV"        : 18.0,   # cP  (Plastic Viscosity)
    "YP"        : 12.0,   # lb/100ft²  (Yield Point)
    # Extra info columns (not parsed — for reference)
    "SurfaceTemp"    : 75.0,    # °F
    "BHTTemp"        : 220.0,   # °F
    "RheologyModel"  : "Bingham Plastic",
    "MudType"        : "Water-Based",
    "SolidsContent"  : 5.0,     # %
    "pH"             : 9.0,
}

# Formations — must cover 0 → TotalDepth with no gaps
# Lithology options: Shale, Sandstone, Limestone, Salt, Dolomite, Anhydrite
FORMATIONS = [
    # ZoneName,      TopDepth, BottomDepth,  PP(ppg),  FG(ppg),  Lithology
    ("Surface",       0.0,     1200.0,        8.6,     10.2,     "Shale"),
    ("Sand A",     1200.0,     3400.0,        9.1,     12.1,     "Sandstone"),
    ("Shale B",    3400.0,     5800.0,        9.8,     13.0,     "Shale"),
    ("Limestone",  5800.0,     7200.0,       10.2,     13.8,     "Limestone"),
    ("Target Sand",7200.0,     9500.0,       11.1,     14.5,     "Sandstone"),
]

# Deviation survey — starts vertical then builds angle
# MD(ft), Inclination(°), Azimuth(°)
SURVEY = [
    (0.0,     0.0,   0.0),
    (1000.0,  0.0,   0.0),
    (2000.0,  0.0,   0.0),
    (3000.0,  5.0,  45.0),
    (4000.0, 15.0,  45.0),
    (5000.0, 28.0,  45.0),
    (6000.0, 38.0,  45.0),
    (7000.0, 45.0,  45.0),
    (8000.0, 45.0,  45.0),
    (9000.0, 45.0,  45.0),
    (9500.0, 45.0,  45.0),
]

# Bit parameters (not parsed by importer — for reference)
BIT = {
    "BitSize"    : 8.5,     # in
    "NozzleCount": 3,
    "Nozzle1"    : 13.0,    # 1/32 in
    "Nozzle2"    : 13.0,
    "Nozzle3"    : 12.0,
}


# ════════════════════════════════════════════════════════════════════════════
# WORKBOOK BUILDER
# ════════════════════════════════════════════════════════════════════════════

wb = Workbook()

# ── Sheet 1: WellInfo ────────────────────────────────────────────────────────
ws = wb.active
ws.title = "WellInfo"
style_sheet(ws)
ws.row_dimensions[1].height = 22
ws.row_dimensions[2].height = 20

# Headers — columns A-D are what the parser reads; E+ are extra info
headers = ["WellName", "TotalDepth (ft)", "CasingOD (in)", "CasingID (in)",
           "DrillPipeOD (in)", "DrillPipeID (in)", "DrillCollarOD (in)", "DrillCollarLength (ft)",
           "BitSize (in)", "NozzleCount", "Nozzle1 (1/32in)", "Nozzle2 (1/32in)", "Nozzle3 (1/32in)"]
write_header_row(ws, headers)

values = [
    WELL_INFO["WellName"],
    WELL_INFO["TotalDepth"],
    WELL_INFO["CasingOD"],
    WELL_INFO["CasingID"],
    DRILL_STRING["DrillPipeOD"],
    DRILL_STRING["DrillPipeID"],
    DRILL_STRING["DrillCollarOD"],
    DRILL_STRING["DrillCollarLength"],
    BIT["BitSize"],
    BIT["NozzleCount"],
    BIT["Nozzle1"],
    BIT["Nozzle2"],
    BIT["Nozzle3"],
]
write_data_row(ws, values, row=2, even=True)

# ── Sheet 2: FluidProps ──────────────────────────────────────────────────────
ws2 = wb.create_sheet("FluidProps")
style_sheet(ws2)
ws2.row_dimensions[1].height = 22
ws2.row_dimensions[2].height = 20

fluid_headers = [
    "MudWeight (ppg)", "FlowRate (gpm)", "PV (cP)", "YP (lb/100ft²)",
    "SurfaceTemp (°F)", "BHT (°F)", "RheologyModel", "MudType",
    "SolidsContent (%)", "pH"
]
write_header_row(ws2, fluid_headers)

fluid_values = [
    FLUID_PROPS["MudWeight"],
    FLUID_PROPS["FlowRate"],
    FLUID_PROPS["PV"],
    FLUID_PROPS["YP"],
    FLUID_PROPS["SurfaceTemp"],
    FLUID_PROPS["BHTTemp"],
    FLUID_PROPS["RheologyModel"],
    FLUID_PROPS["MudType"],
    FLUID_PROPS["SolidsContent"],
    FLUID_PROPS["pH"],
]
write_data_row(ws2, fluid_values, row=2, even=True)

# ── Sheet 3: Formations ──────────────────────────────────────────────────────
ws3 = wb.create_sheet("Formations")
style_sheet(ws3)
ws3.row_dimensions[1].height = 22

form_headers = [
    "ZoneName", "TopDepth (ft)", "BottomDepth (ft)",
    "PP (ppg)", "FG (ppg)", "Lithology"
]
write_header_row(ws3, form_headers)

lith_colors = {
    "Shale"     : "607080",
    "Sandstone" : "D4A843",
    "Limestone" : "D9CDB4",
    "Salt"      : "B0C8D8",
    "Dolomite"  : "A0B4A0",
    "Anhydrite" : "C8B8D0",
}

for i, (name, top, bot, pp, fg, lith) in enumerate(FORMATIONS):
    row_num = i + 2
    ws3.row_dimensions[row_num].height = 18
    values = [name, top, bot, pp, fg, lith]
    for col, val in enumerate(values, start=1):
        cell = ws3.cell(row=row_num, column=col, value=val)
        cell.fill      = row_fill(i % 2 == 0)
        cell.font      = value_font()
        cell.border    = thin_border()
        cell.alignment = center()
    # Colour the lithology cell to match the app's chips
    lith_cell = ws3.cell(row=row_num, column=6)
    lith_cell.fill = PatternFill("solid", fgColor=lith_colors.get(lith, CARD))
    lith_cell.font = Font(color=NAVY if lith in ("Limestone", "Salt") else WHITE,
                          bold=True, name="Calibri", size=10)

# ── Sheet 4: Survey ──────────────────────────────────────────────────────────
ws4 = wb.create_sheet("Survey")
style_sheet(ws4)
ws4.row_dimensions[1].height = 22

survey_headers = ["MD (ft)", "Inclination (°)", "Azimuth (°)"]
write_header_row(ws4, survey_headers)

for i, (md, inc, azi) in enumerate(SURVEY):
    row_num = i + 2
    ws4.row_dimensions[row_num].height = 18
    write_data_row(ws4, [md, inc, azi], row=row_num, even=i % 2 == 0)

# ── Sheet 5: DrillStringSections (reference only) ────────────────────────────
ws5 = wb.create_sheet("DrillStringSections")
style_sheet(ws5)
ws5.row_dimensions[1].height = 22

ds_headers = ["SectionName", "OD (in)", "ID (in)", "Length (ft)", "Weight (lb/ft)"]
write_header_row(ws5, ds_headers)

for i, (name, od, id_, length, wt) in enumerate(DRILL_SECTIONS):
    row_num = i + 2
    ws5.row_dimensions[row_num].height = 18
    write_data_row(ws5, [name, od, id_, length, wt], row=row_num, even=i % 2 == 0)

# ── Final: freeze header rows and save ───────────────────────────────────────
for sheet in wb.worksheets:
    sheet.freeze_panes = "A2"

output_path = "Well_07A_Profile.xlsx"
wb.save(output_path)
print(f"✓ Generated: {output_path}")
print()
print("Sheets written:")
for sheet in wb.worksheets:
    print(f"  • {sheet.title}")
print()
print("Import into WellLogAnalyzer using the 'Import from Excel' button.")
print("The parser reads: WellInfo(A2-D2), FluidProps(A2-D2),")
print("Formations(all rows), Survey(all rows).")
print()
print("── Well Summary ─────────────────────────────────────────────")
print(f"  Well        : {WELL_INFO['WellName']}")
print(f"  Total Depth : {WELL_INFO['TotalDepth']} ft")
print(f"  Casing OD   : {WELL_INFO['CasingOD']} in  /  ID: {WELL_INFO['CasingID']} in")
print(f"  Mud Weight  : {FLUID_PROPS['MudWeight']} ppg")
print(f"  Flow Rate   : {FLUID_PROPS['FlowRate']} gpm")
print(f"  PV / YP     : {FLUID_PROPS['PV']} cP / {FLUID_PROPS['YP']} lb/100ft²")
print(f"  Formations  : {len(FORMATIONS)} zones  (0 → {WELL_INFO['TotalDepth']} ft)")
print(f"  Survey pts  : {len(SURVEY)}  (max inc {max(s[1] for s in SURVEY)}°)")
print("────────────────────────────────────────────────────────────")
